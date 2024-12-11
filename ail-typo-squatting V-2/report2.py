import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Alignment
import os
import sys
import glob

# Check for correct number of arguments
if len(sys.argv) != 3:
    print("Usage: python generate_report.py <output_folder> <report_output_path>")
    sys.exit(1)

# Command-line arguments
output_folder = sys.argv[1]      # Folder containing the resolved domains and screenshots
output_report = sys.argv[2]      # Path for the output .xlsx report file

# Define file paths based on the output folder
resolved_domains_file = os.path.join(output_folder, "resolved.txt")
screenshots_folder = os.path.join(output_folder, "screenshots")

# Validate input paths
if not os.path.exists(output_folder):
    print(f"Output folder not found: {output_folder}")
    sys.exit(1)

if not os.path.exists(resolved_domains_file):
    print(f"Resolved domains file not found: {resolved_domains_file}")
    sys.exit(1)

if not os.path.exists(screenshots_folder):
    print(f"Screenshots folder not found: {screenshots_folder}")
    sys.exit(1)

# Initialize workbook and active sheet
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Resolved Domains Report"

# Define column headers
headers = ["Domain", "Record Type", "IP/CNAME", "HTTP Screenshot", "HTTPS Screenshot"]
for col, header in enumerate(headers, 1):
    ws.cell(row=1, column=col, value=header)
    ws.cell(row=1, column=col).font = Font(bold=True)
    ws.cell(row=1, column=col).alignment = Alignment(horizontal="center")

# Read resolved domains
with open(resolved_domains_file, "r") as f:
    resolved_domains = [line.strip() for line in f if line.strip()]

# Set initial row height to allow room for screenshots
for row in range(2, len(resolved_domains) + 2):
    ws.row_dimensions[row].height = 110  # Increased row height for screenshots

# Define a helper function to find a screenshot file matching the domain
def find_screenshot(domain, protocol, folder):
    pattern = os.path.join(folder, f"{protocol}---{domain}-*.jpeg")  # Match any port number
    matching_files = glob.glob(pattern)
    return matching_files[0] if matching_files else None

# Populate the worksheet with domain data
row = 2  # Start data from the second row
for domain_entry in resolved_domains:
    # Split domain entry to capture domain, record type, and IP/CNAME
    parts = domain_entry.split()
    if len(parts) >= 3:
        domain = parts[0].rstrip('.')  # Strip trailing dot from domain
        record_type = parts[1]
        ip_or_cname = parts[2]

        # Insert domain, record type, and IP/CNAME into the worksheet
        ws.cell(row=row, column=1, value=domain)
        ws.cell(row=row, column=2, value=record_type)
        ws.cell(row=row, column=3, value=ip_or_cname)

        # Find HTTP and HTTPS screenshot files
        http_screenshot_path = find_screenshot(domain, "http", screenshots_folder)
        https_screenshot_path = find_screenshot(domain, "https", screenshots_folder)

        # Initialize image variables for HTTP and HTTPS
        http_img = None
        https_img = None

        # Check for available HTTP screenshot
        if http_screenshot_path and os.path.exists(http_screenshot_path):
            http_img = Image(http_screenshot_path)
            http_img.width, http_img.height = 200, 150  # Increased size for better visibility

        # Check for available HTTPS screenshot
        if https_screenshot_path and os.path.exists(https_screenshot_path):
            https_img = Image(https_screenshot_path)
            https_img.width, https_img.height = 200, 150  # Increased size for better visibility

        # Add HTTP screenshot to the report if available
        if http_img:
            ws.add_image(http_img, f"D{row}")
        else:
            ws.cell(row=row, column=4, value="No HTTP screenshot")

        # Add HTTPS screenshot to the report if available
        if https_img:
            ws.add_image(https_img, f"E{row}")
        else:
            ws.cell(row=row, column=5, value="No HTTPS screenshot")

        row += 1  # Move to the next row

# Adjust column widths
ws.column_dimensions["A"].width = 30  # Domain
ws.column_dimensions["B"].width = 15  # Record Type
ws.column_dimensions["C"].width = 25  # IP/CNAME
ws.column_dimensions["D"].width = 60  # HTTP Screenshot
ws.column_dimensions["E"].width = 60  # HTTPS Screenshot

# Save the workbook
wb.save(output_report)
print(f"Security report saved as {output_report}")
